import io
from datetime import date
from decimal import Decimal
from xml.etree import ElementTree

from openpyxl import load_workbook

from app.export.csv_exporter import export_csv
from app.export.excel_exporter import export_excel
from app.export.service import ExportRow
from app.export.tally_xml_exporter import export_tally_xml


def _row(**overrides) -> ExportRow:
    defaults = dict(
        row_number=1,
        txn_date=date(2026, 1, 15),
        voucher_type="Payment",
        voucher_number="V00001",
        ledger_name="VCT PRODUCTS",
        group_name="Sundry Creditors",
        debit=Decimal("1000.00"),
        credit=Decimal("0"),
        narration="RTGS to VCT PRODUCTS",
    )
    defaults.update(overrides)
    return ExportRow(**defaults)


def test_export_csv_contains_header_and_row():
    content = export_csv([_row()]).decode("utf-8-sig")
    assert "Date,Voucher Type,Voucher Number,Ledger,Group,Debit,Credit,Narration" in content
    assert "2026-01-15" in content
    assert "V00001" in content
    assert "1000.00" in content
    assert "VCT PRODUCTS" in content


def test_export_csv_empty_rows_still_produces_header_only():
    content = export_csv([]).decode("utf-8-sig")
    lines = [line for line in content.splitlines() if line]
    assert len(lines) == 1


def test_export_excel_round_trips_values():
    content = export_excel([_row(), _row(row_number=2, voucher_number="V00002", credit=Decimal("500"), debit=Decimal("0"))])
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active

    header = [cell.value for cell in sheet[1]]
    assert header == ["Date", "Voucher Type", "Voucher Number", "Ledger", "Group", "Debit", "Credit", "Narration"]

    first_data_row = [cell.value for cell in sheet[2]]
    assert first_data_row[2] == "V00001"
    assert first_data_row[5] == 1000.0
    assert first_data_row[6] is None


def test_export_excel_formats_date_column_as_a_date_not_a_raw_serial_number():
    content = export_excel([_row()])
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active

    date_cell = sheet.cell(row=2, column=1)
    # openpyxl round-trips a formatted date cell back as a datetime, not a date -
    # what matters here is the day/month/year and that a date format was applied.
    assert date_cell.value.date() == date(2026, 1, 15)
    assert date_cell.number_format != "General"


def test_export_tally_xml_is_well_formed_and_has_two_ledger_entries_per_voucher():
    content = export_tally_xml([_row()])
    root = ElementTree.fromstring(content)

    vouchers = root.findall(".//VOUCHER")
    assert len(vouchers) == 1
    voucher = vouchers[0]
    assert voucher.get("VCHTYPE") == "Payment"
    assert voucher.findtext("DATE") == "20260115"
    assert voucher.findtext("VOUCHERNUMBER") == "V00001"

    entries = voucher.findall("ALLLEDGERENTRIES.LIST")
    assert len(entries) == 2


def test_export_tally_xml_debit_leg_signs_amounts_correctly():
    # debit=1000: bank leg credited (+1000, No), counterparty leg debited (-1000, Yes)
    content = export_tally_xml([_row(debit=Decimal("1000"), credit=Decimal("0"))])
    root = ElementTree.fromstring(content)
    entries = root.find(".//VOUCHER").findall("ALLLEDGERENTRIES.LIST")

    bank_entry, counterparty_entry = entries
    assert bank_entry.findtext("ISDEEMEDPOSITIVE") == "No"
    assert bank_entry.findtext("AMOUNT") == "1000.00"
    assert counterparty_entry.findtext("LEDGERNAME") == "VCT PRODUCTS"
    assert counterparty_entry.findtext("ISDEEMEDPOSITIVE") == "Yes"
    assert counterparty_entry.findtext("AMOUNT") == "-1000.00"


def test_export_tally_xml_credit_leg_signs_amounts_correctly():
    # credit=500: bank leg debited (-500, Yes), counterparty leg credited (+500, No)
    content = export_tally_xml(
        [_row(voucher_type="Receipt", debit=Decimal("0"), credit=Decimal("500"))]
    )
    root = ElementTree.fromstring(content)
    entries = root.find(".//VOUCHER").findall("ALLLEDGERENTRIES.LIST")

    bank_entry, counterparty_entry = entries
    assert bank_entry.findtext("ISDEEMEDPOSITIVE") == "Yes"
    assert bank_entry.findtext("AMOUNT") == "-500.00"
    assert counterparty_entry.findtext("ISDEEMEDPOSITIVE") == "No"
    assert counterparty_entry.findtext("AMOUNT") == "500.00"


def test_export_tally_xml_escapes_special_characters_in_narration():
    content = export_tally_xml([_row(narration="Tom & Jerry's <Import> Co.")])
    root = ElementTree.fromstring(content)  # would raise ParseError if unescaped
    assert root.find(".//NARRATION").text == "Tom & Jerry's <Import> Co."
